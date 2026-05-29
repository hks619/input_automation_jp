"""Generate the blank multi-sheet .xlsx template.

The template's schema is driven by core.schema so the writer cannot drift
from the parser, splitter, or output writer.
"""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from core.schema import SHEETS, SheetSpec, OUTPUT_HEADER


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
REQUIRED_FILL = PatternFill("solid", fgColor="C6E0B4")
EXAMPLE_FONT = Font(italic=True, color="6F6F6F")


def _write_sheet(wb: Workbook, spec: SheetSpec) -> None:
    ws = wb.create_sheet(spec.name)
    # Header row.
    for col_idx, col_name in enumerate(spec.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        if col_name in spec.required:
            # Mark required column header with a different fill on row 2 (legend row).
            ws.cell(row=2, column=col_idx, value="required").fill = REQUIRED_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col_name) + 2)
    # Example row (row 3).
    if spec.example:
        for col_idx, val in enumerate(spec.example, start=1):
            cell = ws.cell(row=3, column=col_idx, value=val)
            cell.font = EXAMPLE_FONT
    # Notes in a top-right cell (out of the way of data columns).
    notes_col = len(spec.columns) + 2
    ws.cell(row=1, column=notes_col, value="NOTES").font = HEADER_FONT
    ws.cell(row=1, column=notes_col).fill = HEADER_FILL
    ws.cell(row=2, column=notes_col, value=spec.notes).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions[get_column_letter(notes_col)].width = 60
    ws.freeze_panes = "A2"


def _write_instructions(wb: Workbook) -> None:
    ws = wb.create_sheet("Instructions")
    ws["A1"] = "RMS Exposure Preparation Template"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Fill the six data sheets below. Row 2 of each sheet marks required columns."
    ws["A3"] = "Row 3 of each sheet shows an example. Delete it before uploading."
    ws["A4"] = ""

    row = 5
    ws.cell(row=row, column=1, value="Sheet").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Columns").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Required").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Notes").font = Font(bold=True)
    row += 1
    for spec in SHEETS:
        ws.cell(row=row, column=1, value=spec.name)
        ws.cell(row=row, column=2, value=", ".join(spec.columns))
        ws.cell(row=row, column=3, value=", ".join(spec.required))
        ws.cell(row=row, column=4, value=spec.notes).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    row += 2
    ws.cell(row=row, column=1, value="Output (.txt) header — for reference").font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value=", ".join(OUTPUT_HEADER)).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 60


def build_template() -> bytes:
    """Return the .xlsx template as bytes (ready for Streamlit download_button)."""
    wb = Workbook()
    # Drop the default sheet.
    default = wb.active
    wb.remove(default)

    _write_instructions(wb)
    for spec in SHEETS:
        _write_sheet(wb, spec)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_template_to_path(path: str) -> None:
    """Convenience for tests/scripts."""
    with open(path, "wb") as f:
        f.write(build_template())
