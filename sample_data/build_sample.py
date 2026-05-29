"""Generate a filled-in template and the corresponding output .txt.

Run from the repo root:

    python sample_data/build_sample.py

Writes:
    sample_data/sample_filled_template.xlsx
    sample_data/sample_output.txt
    sample_data/sample_loc_cede.csv  (input to the CEDE module)
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.geoprocess import load_postal_lookup, load_secondary_modifier_config
from core.schema import SHEETS, SHEETS_BY_NAME
from core.splitter import run_splitter
from core.txt_writer import build_output_frame, to_tsv_bytes


def _filled_sheets() -> dict:
    account = pd.DataFrame([
        {"ACCNTNUM": "ACC-JP-001", "ACCNTNAME": "Tokyo Marunouchi Office Portfolio",
         "POLICYNUM": "POL-2026-01", "POLICYTYPE": "Treaty", "LOBNAME": "COMM",
         "CEDANTID": "CED-JP-DEMO", "CEDANTNAME": "Demo Cedent JP",
         "PARTOF": 1.0, "PARTOFCUR": "JPY",
         "BLANLIMAMT": 10_000_000_000, "BLANLIMCUR": "JPY"},
        {"ACCNTNUM": "ACC-JP-002", "ACCNTNAME": "Osaka Residential Portfolio",
         "POLICYNUM": "POL-2026-02", "POLICYTYPE": "Treaty", "LOBNAME": "RESID",
         "CEDANTID": "CED-JP-DEMO", "CEDANTNAME": "Demo Cedent JP",
         "PARTOF": 1.0, "PARTOFCUR": "JPY",
         "BLANLIMAMT": 5_000_000_000, "BLANLIMCUR": "JPY"},
    ])
    exp = pd.DataFrame([
        {"ACCNTNUM": "ACC-JP-001", "LOBNAME": "COMM",
         "STATE": "Tokyo", "COUNTY": "Chiyoda", "DISTRICT": "Marunouchi",
         "POSTCODE": "100-0001", "CNTRYCODE": "JP", "CNTRYSCHEME": "ISO", "CRESTA": "JP_13",
         "NUMBLDGS": 50, "NUMFLOORS": 8,
         "BLDG": 3_000_000_000, "CONT": 500_000_000, "BI": 200_000_000,
         "eTIV": 3_700_000_000, "SITELIM": 5_000_000_000, "CURRENCY": "JPY"},
        {"ACCNTNUM": "ACC-JP-002", "LOBNAME": "RESID",
         "STATE": "Osaka", "COUNTY": "Osaka-shi", "DISTRICT": "Kita-ku",
         "POSTCODE": "530-0001", "CNTRYCODE": "JP", "CNTRYSCHEME": "ISO", "CRESTA": "JP_27",
         "NUMBLDGS": 200, "NUMFLOORS": 3,
         "BLDG": 1_500_000_000, "CONT": 300_000_000, "BI": 50_000_000,
         "eTIV": 1_850_000_000, "SITELIM": 2_000_000_000, "CURRENCY": "JPY"},
    ])
    occ = pd.DataFrame([
        {"LOB": "COMM", "LOB_TYPE": "Commercial", "OCC": "Office building",
         "OCCSCHEME": "RMS", "OCCTYPE": "331", "Pre_Rebased_Percentage": 80.0, "Occ_split": None},
        {"LOB": "COMM", "LOB_TYPE": "Commercial", "OCC": "General commercial",
         "OCCSCHEME": "RMS", "OCCTYPE": "320", "Pre_Rebased_Percentage": 20.0, "Occ_split": None},
        {"LOB": "RESID", "LOB_TYPE": "Residential", "OCC": "Apartment",
         "OCCSCHEME": "RMS", "OCCTYPE": "311", "Pre_Rebased_Percentage": 70.0, "Occ_split": None},
        {"LOB": "RESID", "LOB_TYPE": "Residential", "OCC": "Single family dwelling",
         "OCCSCHEME": "RMS", "OCCTYPE": "300", "Pre_Rebased_Percentage": 30.0, "Occ_split": None},
    ])
    cons = pd.DataFrame([
        {"Cedant": "CED-JP-DEMO", "LOBNAME": "COMM", "Construction": "RC moment frame",
         "BLDGSCHEME": "RMS", "BLDGCLASS": "311", "Cons_split": 0.8},
        {"Cedant": "CED-JP-DEMO", "LOBNAME": "COMM", "Construction": "Steel moment frame",
         "BLDGSCHEME": "RMS", "BLDGCLASS": "411", "Cons_split": 0.2},
        {"Cedant": "CED-JP-DEMO", "LOBNAME": "RESID", "Construction": "Reinforced concrete",
         "BLDGSCHEME": "RMS", "BLDGCLASS": "311", "Cons_split": 0.6},
        {"Cedant": "CED-JP-DEMO", "LOBNAME": "RESID", "Construction": "Light wood frame",
         "BLDGSCHEME": "RMS", "BLDGCLASS": "111", "Cons_split": 0.4},
    ])
    bh = pd.DataFrame([
        {"Cedant": "CED-JP-DEMO", "LOB": "COMM", "NUMSTORIES": 8,
         "BLDG_HEIGHT_REF": "Mid-rise (4-9)", "BH_split": 0.7},
        {"Cedant": "CED-JP-DEMO", "LOB": "COMM", "NUMSTORIES": 15,
         "BLDG_HEIGHT_REF": "High-rise (10+)", "BH_split": 0.3},
        {"Cedant": "CED-JP-DEMO", "LOB": "RESID", "NUMSTORIES": 3,
         "BLDG_HEIGHT_REF": "Low-rise (1-3)", "BH_split": 0.6},
        {"Cedant": "CED-JP-DEMO", "LOB": "RESID", "NUMSTORIES": 6,
         "BLDG_HEIGHT_REF": "Mid-rise (4-9)", "BH_split": 0.4},
    ])
    yb = pd.DataFrame([
        {"Cedant": "CED-JP-DEMO", "LOB": "COMM", "YEARBUILT": 2010,
         "YEARBUILT_BAND": "2001-2010", "YB_split": 0.5},
        {"Cedant": "CED-JP-DEMO", "LOB": "COMM", "YEARBUILT": 2018,
         "YEARBUILT_BAND": "2011-2020", "YB_split": 0.5},
        {"Cedant": "CED-JP-DEMO", "LOB": "RESID", "YEARBUILT": 1995,
         "YEARBUILT_BAND": "1991-2000", "YB_split": 0.6},
        {"Cedant": "CED-JP-DEMO", "LOB": "RESID", "YEARBUILT": 2008,
         "YEARBUILT_BAND": "2001-2010", "YB_split": 0.4},
    ])
    return {"Account_Group": account, "EXP_EQ": exp, "Occ": occ,
            "Cons": cons, "BH": bh, "YB": yb}


def _write_filled_xlsx(sheets: dict, path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for spec in SHEETS:
        ws = wb.create_sheet(spec.name)
        for col_idx, col in enumerate(spec.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        df = sheets[spec.name]
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, col in enumerate(spec.columns, start=1):
                if col in df.columns:
                    ws.cell(row=r_idx, column=c_idx, value=getattr(row, col, None))
    wb.save(path)


def main() -> None:
    sheets = _filled_sheets()
    out_dir = ROOT / "sample_data"
    out_dir.mkdir(exist_ok=True)

    _write_filled_xlsx(sheets, out_dir / "sample_filled_template.xlsx")

    disagg, _ = run_splitter(sheets, pruning_threshold=0.0001)
    defaults = {
        "peril_rule": "mirror_eq",
        "default_currency": "JPY",
        "sitelim_rule": "repeat_whole",
        "country_default": "JP",
        "country_scheme_default": "ISO",
    }
    sec_cfg = load_secondary_modifier_config()
    out_df = build_output_frame(disagg, defaults, secondary_rules=sec_cfg.get("rules", []))
    txt = to_tsv_bytes(out_df, line_ending="\r\n")
    (out_dir / "sample_output.txt").write_bytes(txt)

    # CEDE sample.
    loc_cede = pd.DataFrame([
        {"LocID": 1, "PortNum": 1, "AccGrpID": 1, "DedType1": "FR",
         "DedAmt1": 1_000_000, "DedCur1": "JPY", "DedTxt1": ""},
        {"LocID": 2, "PortNum": 1, "AccGrpID": 1, "DedType1": "S",
         "DedAmt1": 0, "DedCur1": "JPY", "DedTxt1": ""},
        {"LocID": 3, "PortNum": 1, "AccGrpID": 1, "DedType1": "FR",
         "DedAmt1": 500_000, "DedCur1": "JPY", "DedTxt1": "legacy-note"},
    ])
    loc_cede.to_csv(out_dir / "sample_loc_cede.csv", index=False)

    print(f"Wrote: {out_dir/'sample_filled_template.xlsx'}")
    print(f"Wrote: {out_dir/'sample_output.txt'} ({len(out_df)} rows)")
    print(f"Wrote: {out_dir/'sample_loc_cede.csv'} ({len(loc_cede)} rows)")


if __name__ == "__main__":
    main()
