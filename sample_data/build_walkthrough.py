"""Build a walkthrough dummy template + matching CEDE CSV.

Produces a small, easy-to-trace dataset so an analyst can step through every
page of the tool and see exactly how the txt file is generated.

Inputs (in EXP_EQ):
    ACC-DEMO-001 / LOB=RESID
        Tokyo Marunouchi, 100 buildings,
        BLDG=1,000,000,000, CONT=200,000,000, BI=50,000,000
        SITELIM=2,000,000,000

Splits (all for LOB=RESID):
    Occ:  Apartment 60% (OCCTYPE 311) | SFD 40% (OCCTYPE 300)
    Cons: RC 70% (BLDGCLASS 311)      | Wood 30% (BLDGCLASS 100)
    BH:   3 storeys 60%               | 8 storeys 40%
    YB:   2005 50%                    | 1985 50%

Cross-product: 2 × 2 × 2 × 2 = 16 output rows.

For one of those rows you can hand-check the math:
    Apartment × RC × 3-stry × 2005
    proportion = 0.6 × 0.7 × 0.6 × 0.5 = 0.126
    EQCV1VAL (BLDG) = 1,000,000,000 × 0.126 = 126,000,000
    EQCV2VAL (CONT) =   200,000,000 × 0.126 =  25,200,000
    EQCV3VAL (BI)   =    50,000,000 × 0.126 =   6,300,000
    NUMBLDGS        = round(100 × 0.126)    = 13  (sum of all 16 rows = 100)

Outputs written:
    sample_data/walkthrough_filled_template.xlsx
    sample_data/walkthrough_expected_output.txt
    sample_data/walkthrough_loc_cede.csv
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.geoprocess import load_secondary_modifier_config
from core.schema import SHEETS
from core.splitter import run_splitter
from core.txt_writer import build_output_frame, to_tsv_bytes


def _filled_sheets() -> dict:
    account = pd.DataFrame([{
        "ACCNTNUM": "ACC-DEMO-001",
        "ACCNTNAME": "Walkthrough Account",
        "POLICYNUM": "POL-2026-DEMO",
        "POLICYTYPE": "Treaty",
        "LOBNAME": "RESID",
        "CEDANTID": "CED-DEMO",
        "CEDANTNAME": "Demo Cedent",
        "PARTOF": 1.0,
        "PARTOFCUR": "JPY",
        "BLANLIMAMT": 5_000_000_000,
        "BLANLIMCUR": "JPY",
    }])

    exp = pd.DataFrame([{
        "ACCNTNUM": "ACC-DEMO-001",
        "LOBNAME": "RESID",
        "STATE": "Tokyo",
        "COUNTY": "Chiyoda",
        "DISTRICT": "Marunouchi",
        "POSTCODE": "100-0001",
        "CNTRYCODE": "JP",
        "CNTRYSCHEME": "ISO",
        "CRESTA": "JP_13",
        "NUMBLDGS": 100,
        "NUMFLOORS": 3,
        "BLDG": 1_000_000_000,
        "CONT":   200_000_000,
        "BI":      50_000_000,
        "eTIV": 1_250_000_000,
        "SITELIM": 2_000_000_000,
        "CURRENCY": "JPY",
    }])

    occ = pd.DataFrame([
        {"LOB": "RESID", "LOB_TYPE": "Residential",
         "OCC": "Apartment", "OCCSCHEME": "RMS", "OCCTYPE": "311",
         "Pre_Rebased_Percentage": 60.0, "Occ_split": None},
        {"LOB": "RESID", "LOB_TYPE": "Residential",
         "OCC": "Single family dwelling", "OCCSCHEME": "RMS", "OCCTYPE": "300",
         "Pre_Rebased_Percentage": 40.0, "Occ_split": None},
    ])

    cons = pd.DataFrame([
        {"Cedant": "CED-DEMO", "LOBNAME": "RESID",
         "Construction": "Reinforced concrete",
         "BLDGSCHEME": "RMS", "BLDGCLASS": "311", "Cons_split": 0.7},
        {"Cedant": "CED-DEMO", "LOBNAME": "RESID",
         "Construction": "Wood frame",
         "BLDGSCHEME": "RMS", "BLDGCLASS": "100", "Cons_split": 0.3},
    ])

    bh = pd.DataFrame([
        {"Cedant": "CED-DEMO", "LOB": "RESID",
         "NUMSTORIES": 3, "BLDG_HEIGHT_REF": "Low-rise (1-3)", "BH_split": 0.6},
        {"Cedant": "CED-DEMO", "LOB": "RESID",
         "NUMSTORIES": 8, "BLDG_HEIGHT_REF": "Mid-rise (4-9)", "BH_split": 0.4},
    ])

    yb = pd.DataFrame([
        {"Cedant": "CED-DEMO", "LOB": "RESID",
         "YEARBUILT": 2005, "YEARBUILT_BAND": "2001-2010", "YB_split": 0.5},
        {"Cedant": "CED-DEMO", "LOB": "RESID",
         "YEARBUILT": 1985, "YEARBUILT_BAND": "1981-1990", "YB_split": 0.5},
    ])
    return {"Account_Group": account, "EXP_EQ": exp, "Occ": occ,
            "Cons": cons, "BH": bh, "YB": yb}


def _write_filled_xlsx(sheets: dict, path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for spec in SHEETS:
        ws = wb.create_sheet(spec.name)
        for col_idx, col in enumerate(spec.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
        df = sheets[spec.name]
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, col in enumerate(spec.columns, start=1):
                if col in df.columns:
                    ws.cell(row=r_idx, column=c_idx, value=getattr(row, col, None))
    wb.save(path)


def main() -> None:
    sheets = _filled_sheets()
    out_dir = ROOT / "sample_data"
    out_dir.mkdir(exist_ok=True)

    _write_filled_xlsx(sheets, out_dir / "walkthrough_filled_template.xlsx")

    disagg, _ = run_splitter(sheets, pruning_threshold=0.0)
    defaults = {
        "peril_rule": "mirror_eq",
        "default_currency": "JPY",
        "sitelim_rule": "repeat_whole",
        "country_default": "JP",
        "country_scheme_default": "ISO",
    }
    sec_cfg = load_secondary_modifier_config()
    out_df = build_output_frame(disagg, defaults, secondary_rules=sec_cfg.get("rules", []))
    (out_dir / "walkthrough_expected_output.txt").write_bytes(
        to_tsv_bytes(out_df, line_ending="\r\n")
    )

    # CEDE walkthrough: 4 rows, 2 of which are FR.
    loc_cede = pd.DataFrame([
        {"LocID": 101, "PortNum": 1, "AccGrpID": 1, "DedType1": "FR",
         "DedAmt1": 1_000_000, "DedCur1": "JPY", "DedTxt1": ""},
        {"LocID": 102, "PortNum": 1, "AccGrpID": 1, "DedType1": "S",
         "DedAmt1": 500_000, "DedCur1": "JPY", "DedTxt1": ""},
        {"LocID": 103, "PortNum": 1, "AccGrpID": 1, "DedType1": "FR",
         "DedAmt1": 2_000_000, "DedCur1": "JPY", "DedTxt1": "legacy"},
        {"LocID": 104, "PortNum": 1, "AccGrpID": 1, "DedType1": "S",
         "DedAmt1": 0, "DedCur1": "JPY", "DedTxt1": ""},
    ])
    loc_cede.to_csv(out_dir / "walkthrough_loc_cede.csv", index=False)

    # Print summary so analyst can verify by eye.
    print(f"Wrote: {out_dir/'walkthrough_filled_template.xlsx'}")
    print(f"Wrote: {out_dir/'walkthrough_expected_output.txt'} ({len(out_df)} rows)")
    print(f"Wrote: {out_dir/'walkthrough_loc_cede.csv'} ({len(loc_cede)} rows; 2 FR rows)")
    print()
    print("Input totals (EXP_EQ):")
    e = sheets["EXP_EQ"].iloc[0]
    print(f"  NUMBLDGS = {int(e['NUMBLDGS'])}")
    print(f"  BLDG     = {int(e['BLDG']):,}")
    print(f"  CONT     = {int(e['CONT']):,}")
    print(f"  BI       = {int(e['BI']):,}")
    print()
    print("Output totals (after splitter):")
    print(f"  Σ NUMBLDGS = {int(out_df['NUMBLDGS'].sum())}")
    print(f"  Σ EQCV1VAL = {int(out_df['EQCV1VAL'].sum()):,}")
    print(f"  Σ EQCV2VAL = {int(out_df['EQCV2VAL'].sum()):,}")
    print(f"  Σ EQCV3VAL = {int(out_df['EQCV3VAL'].sum()):,}")


if __name__ == "__main__":
    main()
